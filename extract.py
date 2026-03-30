from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "required_columns_output"
MISSING_CODE_TEXT = "code is not mention"


def read_ekcc_csv(file_path: Path) -> pd.DataFrame:
    return pd.read_csv(file_path, encoding="latin1", low_memory=False)


def read_lgd_workbook(file_path: Path, required_columns: list[str]) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(
        BytesIO(file_path.read_bytes()),
        read_only=True,
        data_only=True,
    )
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = next(
        worksheet.iter_rows(min_row=2, max_row=2, values_only=True)
    )
    header_index_map = {str(value).strip(): index for index, value in enumerate(header_row) if value is not None}

    missing_headers = [column for column in required_columns if column not in header_index_map]
    if missing_headers:
        raise KeyError(f"Missing LGD columns in {file_path.name}: {missing_headers}")

    extracted_rows: list[dict[str, object]] = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        extracted_rows.append(
            {
                column: row[header_index_map[column]]
                for column in required_columns
            }
        )

    return pd.DataFrame(extracted_rows, columns=required_columns)


def normalize_code(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    stringified = numeric.astype("Int64").astype("string")
    fallback = series.astype("string").str.strip()
    result = stringified.fillna(fallback)
    result = result.replace({"<NA>": pd.NA, "nan": pd.NA, "None": pd.NA, "": pd.NA})
    return result


def fill_missing_code(series: pd.Series) -> pd.Series:
    cleaned = series.astype("string").str.strip()
    cleaned = cleaned.replace({"<NA>": pd.NA, "nan": pd.NA, "None": pd.NA, "": pd.NA})
    return cleaned.fillna(MISSING_CODE_TEXT)


def ensure_output_dirs(output_dir: Path) -> tuple[Path, Path]:
    ekcc_dir = output_dir / "ekcc"
    lgd_dir = output_dir / "lgd"
    ekcc_dir.mkdir(parents=True, exist_ok=True)
    lgd_dir.mkdir(parents=True, exist_ok=True)
    return ekcc_dir, lgd_dir


def file_exists(file_path: Path) -> bool:
    return file_path.exists() and file_path.is_file()


def build_ekcc_outputs(data_dir: Path) -> dict[str, pd.DataFrame]:
    ekcc_dir = data_dir / "ekcc"
    state_path = ekcc_dir / "ekcc_state.csv"
    district_path = ekcc_dir / "ekcc_district.csv"
    subdistrict_path = ekcc_dir / "ekcc_subdistrict.csv"
    village_path = ekcc_dir / "ekcc_village.csv"

    outputs: dict[str, pd.DataFrame] = {}
    state_lookup: pd.DataFrame | None = None
    district_lookup: pd.DataFrame | None = None
    subdistrict_df: pd.DataFrame | None = None

    if file_exists(state_path):
        state_raw = read_ekcc_csv(state_path)
        state_lookup = state_raw.loc[:, ["id", "state_name", "state_code", "state_type"]].copy()
        state_lookup["state_id_key"] = normalize_code(state_lookup["id"])
        state_lookup["state_code"] = fill_missing_code(normalize_code(state_lookup["state_code"]))
        state_lookup = state_lookup.rename(columns={"state_name": "State Name", "state_type": "State Type"})
        outputs["state"] = state_lookup.loc[:, ["State Name", "state_code", "State Type"]].rename(
            columns={"state_code": "State Code"}
        )

    if file_exists(district_path):
        district_raw = read_ekcc_csv(district_path)
        district_lookup = district_raw.loc[:, ["id", "district_name", "distrcit_code", "state_master_id"]].copy()
        district_lookup["district_id_key"] = normalize_code(district_lookup["id"])
        district_lookup["District Code"] = fill_missing_code(normalize_code(district_lookup["distrcit_code"]))
        district_lookup["state_id_key"] = normalize_code(district_lookup["state_master_id"])
        if state_lookup is not None:
            district_lookup = district_lookup.merge(
                state_lookup.loc[:, ["state_id_key", "State Name", "state_code"]],
                on="state_id_key",
                how="left",
            )
            district_lookup["state_code"] = fill_missing_code(district_lookup["state_code"])
            district_lookup["State Name"] = district_lookup["State Name"].astype("string").fillna("")
        else:
            district_lookup["state_code"] = MISSING_CODE_TEXT
            district_lookup["State Name"] = ""
        district_lookup = district_lookup.rename(columns={"district_name": "District Name", "state_code": "State Code"})
        outputs["district"] = district_lookup.loc[:, ["State Code", "State Name", "District Code", "District Name"]].copy()

    if file_exists(subdistrict_path):
        subdistrict_raw = read_ekcc_csv(subdistrict_path)
        subdistrict_lookup = subdistrict_raw.loc[
            :, ["sub_district_name", "sub_district_code", "district_master_id"]
        ].copy()
        subdistrict_lookup["district_id_key"] = normalize_code(subdistrict_lookup["district_master_id"])
        subdistrict_lookup["Sub District Code"] = fill_missing_code(
            normalize_code(subdistrict_lookup["sub_district_code"])
        )
        if district_lookup is not None:
            subdistrict_lookup = subdistrict_lookup.merge(
                district_lookup.loc[:, ["district_id_key", "District Code", "District Name", "State Name"]],
                on="district_id_key",
                how="left",
            )
            subdistrict_lookup["District Code"] = fill_missing_code(subdistrict_lookup["District Code"])
            subdistrict_lookup["District Name"] = subdistrict_lookup["District Name"].astype("string").fillna("")
            subdistrict_lookup["State Name"] = subdistrict_lookup["State Name"].astype("string").fillna("")
        else:
            subdistrict_lookup["District Code"] = MISSING_CODE_TEXT
            subdistrict_lookup["District Name"] = ""
            subdistrict_lookup["State Name"] = ""
        subdistrict_lookup = subdistrict_lookup.rename(columns={"sub_district_name": "Sub District Name"})
        subdistrict_df = subdistrict_lookup.loc[
            :, ["Sub District Name", "Sub District Code", "District Code", "District Name", "State Name"]
        ].copy()
        outputs["subdistrict"] = subdistrict_df

    if file_exists(village_path):
        village_raw = read_ekcc_csv(village_path)
        village_df = village_raw.loc[:, ["village_name", "village_code", "subdistrict_code"]].copy()
        village_df["Village Code"] = fill_missing_code(normalize_code(village_df["village_code"]))
        village_df["Village Census Code"] = fill_missing_code(normalize_code(village_raw["village_census_code"]))
        village_df["Sub District Code"] = fill_missing_code(normalize_code(village_df["subdistrict_code"]))
        if subdistrict_df is not None:
            subdistrict_parent = subdistrict_df.loc[
                :, ["Sub District Code", "Sub District Name", "District Code", "District Name", "State Name"]
            ].drop_duplicates(subset=["Sub District Code"])
            village_df = village_df.merge(subdistrict_parent, on="Sub District Code", how="left")
            village_df["Sub District Name"] = village_df["Sub District Name"].astype("string").fillna("")
            village_df["District Code"] = fill_missing_code(village_df["District Code"])
            village_df["District Name"] = village_df["District Name"].astype("string").fillna("")
            village_df["State Name"] = village_df["State Name"].astype("string").fillna("")
        else:
            village_df["Sub District Name"] = ""
            village_df["District Code"] = MISSING_CODE_TEXT
            village_df["District Name"] = ""
            village_df["State Name"] = ""
        village_df = village_df.rename(columns={"village_name": "Village Name"})
        outputs["village"] = village_df.loc[
            :,
            [
                "Village Name",
                "Village Code",
                "Village Census Code",
                "Sub District Code",
                "Sub District Name",
                "District Code",
                "District Name",
                "State Name",
            ],
        ].copy()

    return outputs


def build_lgd_outputs(data_dir: Path) -> dict[str, pd.DataFrame]:
    lgd_dir = data_dir / "lgd"
    state_path = lgd_dir / "lgd_state.csv"
    district_path = lgd_dir / "lgd_district.csv"
    subdistrict_path = lgd_dir / "lgd_subdistrict.csv"
    village_path = lgd_dir / "lgd_village.csv"

    outputs: dict[str, pd.DataFrame] = {}
    state_lookup: pd.DataFrame | None = None
    district_lookup: pd.DataFrame | None = None
    subdistrict_lookup: pd.DataFrame | None = None

    if file_exists(state_path):
        state_raw = read_lgd_workbook(
            state_path,
            ["State Code", "State Name (In English)", "State or UT"],
        )
        state_df = pd.DataFrame(
            {
                "State Name": state_raw["State Name (In English)"].astype("string"),
                "State Code": fill_missing_code(normalize_code(state_raw["State Code"])),
                "State Type": state_raw["State or UT"].astype("string"),
            }
        )
        outputs["state"] = state_df
        state_lookup = state_df.drop_duplicates(subset=["State Code"]).copy()

    if file_exists(district_path):
        district_raw = read_lgd_workbook(
            district_path,
            ["State Code", "District Code", "District Name(In English)"],
        )
        district_df = pd.DataFrame(
            {
                "State Code": fill_missing_code(normalize_code(district_raw["State Code"])),
                "District Code": fill_missing_code(normalize_code(district_raw["District Code"])),
                "District Name": district_raw["District Name(In English)"].astype("string"),
            }
        )
        if state_lookup is not None:
            district_df = district_df.merge(
                state_lookup.loc[:, ["State Code", "State Name"]],
                on="State Code",
                how="left",
            )
            district_df["State Name"] = district_df["State Name"].astype("string").fillna("")
        else:
            district_df["State Name"] = ""
        district_df = district_df.loc[:, ["State Code", "State Name", "District Code", "District Name"]]
        outputs["district"] = district_df
        district_lookup = district_df.drop_duplicates(subset=["District Code"]).copy()

    if file_exists(subdistrict_path):
        subdistrict_raw = read_lgd_workbook(
            subdistrict_path,
            ["District Code", "Sub-district Code", "Sub-district Name"],
        )
        subdistrict_df = pd.DataFrame(
            {
                "Sub District Name": subdistrict_raw["Sub-district Name"].astype("string"),
                "Sub District Code": fill_missing_code(normalize_code(subdistrict_raw["Sub-district Code"])),
                "District Code": fill_missing_code(normalize_code(subdistrict_raw["District Code"])),
            }
        )
        if district_lookup is not None:
            subdistrict_df = subdistrict_df.merge(
                district_lookup.loc[:, ["District Code", "District Name", "State Name"]],
                on="District Code",
                how="left",
            )
            subdistrict_df["District Name"] = subdistrict_df["District Name"].astype("string").fillna("")
            subdistrict_df["State Name"] = subdistrict_df["State Name"].astype("string").fillna("")
        else:
            subdistrict_df["District Name"] = ""
            subdistrict_df["State Name"] = ""
        subdistrict_df = subdistrict_df.loc[
            :, ["Sub District Name", "Sub District Code", "District Code", "District Name", "State Name"]
        ]
        outputs["subdistrict"] = subdistrict_df
        subdistrict_lookup = subdistrict_df.drop_duplicates(subset=["Sub District Code"]).copy()

    if file_exists(village_path):
        village_raw = read_lgd_workbook(
            village_path,
            ["Village Name (In English)", "Village Code", "Sub-District Code", "Census 2001 Code", "Census 2011 Code"],
        )
        village_df = pd.DataFrame(
            {
                "Village Name": village_raw["Village Name (In English)"].astype("string"),
                "Village Code": fill_missing_code(normalize_code(village_raw["Village Code"])),
                "Village Census Code": fill_missing_code(
                    normalize_code(village_raw["Census 2011 Code"]).where(
                        pd.to_numeric(village_raw["Census 2011 Code"], errors="coerce").notna(),
                        normalize_code(village_raw["Census 2001 Code"]),
                    )
                ),
                "Sub District Code": fill_missing_code(normalize_code(village_raw["Sub-District Code"])),
            }
        )
        if subdistrict_lookup is not None:
            village_df = village_df.merge(
                subdistrict_lookup.loc[
                    :, ["Sub District Code", "Sub District Name", "District Code", "District Name", "State Name"]
                ],
                on="Sub District Code",
                how="left",
            )
            village_df["Sub District Name"] = village_df["Sub District Name"].astype("string").fillna("")
            village_df["District Code"] = fill_missing_code(village_df["District Code"])
            village_df["District Name"] = village_df["District Name"].astype("string").fillna("")
            village_df["State Name"] = village_df["State Name"].astype("string").fillna("")
        else:
            village_df["Sub District Name"] = ""
            village_df["District Code"] = MISSING_CODE_TEXT
            village_df["District Name"] = ""
            village_df["State Name"] = ""
        village_df = village_df.loc[
            :,
            [
                "Village Name",
                "Village Code",
                "Village Census Code",
                "Sub District Code",
                "Sub District Name",
                "District Code",
                "District Name",
                "State Name",
            ],
        ]
        outputs["village"] = village_df

    return outputs


def write_outputs(dataframes: dict[str, pd.DataFrame], output_dir: Path) -> None:
    for file_stem, dataframe in dataframes.items():
        dataframe.to_csv(output_dir / f"{file_stem}.csv", index=False)


def run_extraction(data_dir: Path = DATA_DIR, output_dir: Path = OUTPUT_DIR) -> dict[str, dict[str, pd.DataFrame]]:
    ekcc_dir, lgd_dir = ensure_output_dirs(output_dir)
    ekcc_outputs = build_ekcc_outputs(data_dir)
    lgd_outputs = build_lgd_outputs(data_dir)

    write_outputs(ekcc_outputs, ekcc_dir)
    write_outputs(lgd_outputs, lgd_dir)

    return {"ekcc": ekcc_outputs, "lgd": lgd_outputs}


def main() -> None:
    outputs = run_extraction(DATA_DIR, OUTPUT_DIR)

    print(f"Generated EKCC files in: {OUTPUT_DIR / 'ekcc'}")
    print(f"Generated LGD files in: {OUTPUT_DIR / 'lgd'}")
    for source_name, source_outputs in outputs.items():
        for file_stem, dataframe in source_outputs.items():
            print(f"{source_name}/{file_stem}.csv -> {dataframe.shape}")


if __name__ == "__main__":
    main()
