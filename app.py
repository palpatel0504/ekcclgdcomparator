from __future__ import annotations

from io import BytesIO
from pathlib import Path
import shutil
import zipfile

import openpyxl
import pandas as pd
import streamlit as st

from compare import write_comparison_files
from extract import run_extraction


BASE_DIR = Path(__file__).resolve().parent
SESSION_ROOT = BASE_DIR / "streamlit_workspace"
UPLOAD_DATA_DIR = SESSION_ROOT / "data"
REQUIRED_OUTPUT_DIR = SESSION_ROOT / "required_columns_output"
COMPARISON_OUTPUT_DIR = SESSION_ROOT / "comparison_output"

EXPECTED_UPLOADS = {
    "EKCC State": ("ekcc", "ekcc_state.csv"),
    "EKCC District": ("ekcc", "ekcc_district.csv"),
    "EKCC Subdistrict": ("ekcc", "ekcc_subdistrict.csv"),
    "EKCC Village": ("ekcc", "ekcc_village.csv"),
    "LGD State": ("lgd", "lgd_state.csv"),
    "LGD District": ("lgd", "lgd_district.csv"),
    "LGD Subdistrict": ("lgd", "lgd_subdistrict.csv"),
    "LGD Village": ("lgd", "lgd_village.csv"),
}


st.set_page_config(page_title="EKCC vs LGD Comparator", layout="wide")


def reset_workspace() -> None:
    if SESSION_ROOT.exists():
        shutil.rmtree(SESSION_ROOT)
    (UPLOAD_DATA_DIR / "ekcc").mkdir(parents=True, exist_ok=True)
    (UPLOAD_DATA_DIR / "lgd").mkdir(parents=True, exist_ok=True)


def save_uploaded_files(uploaded_files: dict[str, object]) -> None:
    reset_workspace()
    for label, uploaded_file in uploaded_files.items():
        if uploaded_file is None:
            continue
        folder_name, target_name = EXPECTED_UPLOADS[label]
        target_path = UPLOAD_DATA_DIR / folder_name / target_name
        target_path.write_bytes(uploaded_file.getbuffer())


def list_csv_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    return sorted(directory.rglob("*.csv"))


def read_preview(file_path: Path, row_count: int) -> pd.DataFrame:
    try:
        return pd.read_csv(file_path, nrows=row_count, dtype="string")
    except Exception:
        workbook = openpyxl.load_workbook(BytesIO(file_path.read_bytes()), read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(min_row=1, max_row=row_count, values_only=True))
        return pd.DataFrame(rows)


def has_comparable_outputs() -> bool:
    entities = ("state", "district", "subdistrict", "village")
    for entity in entities:
        if (REQUIRED_OUTPUT_DIR / "ekcc" / f"{entity}.csv").exists() and (REQUIRED_OUTPUT_DIR / "lgd" / f"{entity}.csv").exists():
            return True
    return False


def zip_directory(directory: Path) -> bytes:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_path in sorted(directory.rglob("*")):
            if file_path.is_file():
                archive.write(file_path, arcname=file_path.relative_to(directory))
    buffer.seek(0)
    return buffer.getvalue()


st.title("EKCC and LGD Comparison")
st.caption("Upload the source files, generate the outputs, and preview the CSVs.")

with st.sidebar:
    st.header("Upload Files")
    uploaded_files: dict[str, object] = {}
    for label in EXPECTED_UPLOADS:
        uploaded_files[label] = st.file_uploader(
            label,
            type=["csv", "xlsx"],
            key=label,
        )
    if st.button("Save Uploaded Files", width="stretch", disabled=not any(uploaded_files.values())):
        save_uploaded_files(uploaded_files)
        st.success("Uploaded files saved.")

    st.divider()
    if st.button("Generate Required Columns", width="stretch", disabled=not any(list_csv_files(UPLOAD_DATA_DIR))):
        with st.spinner("Generating required columns output..."):
            run_extraction(UPLOAD_DATA_DIR, REQUIRED_OUTPUT_DIR)
        st.success("Required columns output generated.")

    if st.button(
        "Generate Comparison Files",
        width="stretch",
        disabled=not has_comparable_outputs(),
    ):
        with st.spinner("Generating comparison output..."):
            write_comparison_files(REQUIRED_OUTPUT_DIR, COMPARISON_OUTPUT_DIR)
        st.success("Comparison output generated.")

    st.divider()
    preview_rows = st.slider("Preview rows", min_value=5, max_value=100, value=20, step=5)


uploaded_col, required_col, comparison_col = st.columns(3)
uploaded_col.metric("Uploaded Source Files", len(list_csv_files(UPLOAD_DATA_DIR)))
required_col.metric("Required Output Files", len(list_csv_files(REQUIRED_OUTPUT_DIR)))
comparison_col.metric("Comparison Files", len(list_csv_files(COMPARISON_OUTPUT_DIR)))

tab_uploads, tab_required, tab_comparison = st.tabs(
    ["Uploaded Files", "Required Columns Output", "Comparison Output"]
)

with tab_uploads:
    uploaded_source_files = list_csv_files(UPLOAD_DATA_DIR)
    if not uploaded_source_files:
        st.info("Upload any EKCC/LGD files you want to work with, then click Save Uploaded Files.")
    else:
        selected_upload = st.selectbox(
            "Choose an uploaded source file",
            options=uploaded_source_files,
            format_func=lambda path: str(path.relative_to(SESSION_ROOT)),
            key="uploaded_file",
        )
        st.dataframe(read_preview(selected_upload, preview_rows), width="stretch")

with tab_required:
    required_files = list_csv_files(REQUIRED_OUTPUT_DIR)
    if not required_files:
        st.info("No required output files found yet. Generate them from the sidebar.")
    else:
        selected_required = st.selectbox(
            "Choose a required output file",
            options=required_files,
            format_func=lambda path: str(path.relative_to(SESSION_ROOT)),
            key="required_file",
        )
        st.dataframe(read_preview(selected_required, preview_rows), width="stretch")
        st.download_button(
            "Download Selected Required Output",
            data=selected_required.read_bytes(),
            file_name=selected_required.name,
            mime="text/csv",
            width="stretch",
        )
        st.download_button(
            "Download All Required Output",
            data=zip_directory(REQUIRED_OUTPUT_DIR),
            file_name="required_columns_output.zip",
            mime="application/zip",
            width="stretch",
        )

with tab_comparison:
    comparison_files = list_csv_files(COMPARISON_OUTPUT_DIR)
    if not comparison_files:
        st.info("No comparison files found yet. Generate them from the sidebar.")
    else:
        selected_comparison = st.selectbox(
            "Choose a comparison file",
            options=comparison_files,
            format_func=lambda path: str(path.relative_to(SESSION_ROOT)),
            key="comparison_file",
        )
        st.dataframe(read_preview(selected_comparison, preview_rows), width="stretch")
        st.download_button(
            "Download Selected Comparison File",
            data=selected_comparison.read_bytes(),
            file_name=selected_comparison.name,
            mime="text/csv",
            width="stretch",
        )
        st.download_button(
            "Download All Comparison Output",
            data=zip_directory(COMPARISON_OUTPUT_DIR),
            file_name="comparison_output.zip",
            mime="application/zip",
            width="stretch",
        )


# with st.expander("How To Run"):
#     st.code("streamlit run /Users/palrpatel/Desktop/FINALFIX/app.py", language="bash")
